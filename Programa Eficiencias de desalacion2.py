#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Programa Eficiencias de desalacion - Versión con fórmulas Excel (español)
- Lee Excel sin cabeceras fijas
- Reconstruye columna Tiempo desde la 2ª columna del Excel (índice 1)
- Usuario elige variable base; se detectan columnas por desalador
- Crea hojas por desalador:
    - Valores_mayor_crit (filtrado por base > crit)
    - Valores_menor_igual_crit (filtrado por base <= crit)
    - Todos_los_valores (sin filtrar)
    - Resumen_Estadistico (con fórmulas Excel en español usando Opción 1)
- Genera gráficos (opcional) y los inserta en un .xlsx por desalador
"""

import os
import sys
import re
import io
import unicodedata
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as xlImage


# -------------------------
# CONFIG
# -------------------------
RUTA_EXCEL = "PI Tarragona Desaladores definitivo.xlsx"  # <- Cambia si hace falta
DIRECTORIO_SALIDA = "Resultados_Desalacion"
if not os.path.exists(DIRECTORIO_SALIDA):
    os.makedirs(DIRECTORIO_SALIDA)

# -------------------------
# UTILIDADES
# -------------------------
def normalizar(txt):
    if txt is None:
        return ""
    txt = str(txt).strip().lower()
    txt = ''.join(c for c in unicodedata.normalize('NFD', txt) if unicodedata.category(c) != 'Mn')
    return txt

def clean_token(s):
    if s is None:
        return ""
    s = str(s)
    s = ''.join(ch for ch in unicodedata.normalize('NFD', s) if unicodedata.category(ch) != 'Mn')
    s = re.sub(r"[^A-Za-z0-9]", "", s)
    return s.lower()

def insertar_imagen_ws(ws, buf, posicion):
    img = xlImage(buf)
    img.anchor = posicion
    ws.add_image(img)

# -------------------------
# LECTURA DEL EXCEL SIN CABECERAS
# -------------------------
def leer_hoja_sin_encabezado(ruta_excel, nombre_hoja):
    try:
        df_raw = pd.read_excel(ruta_excel, sheet_name=nombre_hoja, header=None, engine="openpyxl")
    except Exception as e:
        print("❌ Error leyendo archivo/excel:", e)
        sys.exit(1)
    return df_raw

# -------------------------
# DETECTAR FILA DE INICIO DE DATOS
# -------------------------
def detectar_fila_inicio_datos(df_raw):
    """
    Detecta la fila en la que realmente empiezan los datos,
    ignorando encabezados, medias, desviaciones, unidades, etc.
    Funciona para cualquier hoja.
    """

    palabras_ruido = [
        "media", "desviacion", "max", "min",
        "servidor", "unidades", "escala", "ph",
        "tension", "consumo", "eficiencia"
    ]

    nfilas, ncols = df_raw.shape

    for i in range(nfilas):
        fila = df_raw.iloc[i, :]

        # --- 1) Si la fila contiene texto de encabezado → descartar ---
        texto_fila = " ".join(str(x).lower() for x in fila if pd.notna(x))
        if any(p in texto_fila for p in palabras_ruido):
            continue

        # --- 2) Contar cuántas columnas parecen numéricas ---
        num_ok = 0
        date_ok = 0

        for v in fila:
            if pd.isna(v):
                continue

            # A) ¿Es numérico?
            try:
                float(str(v).replace(",", "."))
                num_ok += 1
                continue
            except:
                pass

            # B) ¿Es fecha?
            if isinstance(v, (pd.Timestamp,)):
                date_ok += 1
                continue

            try:
                pd.to_datetime(v, errors="raise")
                date_ok += 1
            except:
                pass

        # --- 3) Condición para considerar que es fila de datos ---
        # Regla flexible:
        # - la columna 1 es fecha O
        # - más del 40% de columnas son numéricas O
        # - hay al menos 3 columnas numéricas
        col1 = fila.iloc[1] if len(fila) > 1 else None
        col1_es_fecha = False
        try:
            pd.to_datetime(col1, errors="raise")
            col1_es_fecha = True
        except:
            pass

        if col1_es_fecha:
            return i

        if num_ok >= max(3, int(ncols * 0.40)):
            return i

    # Si no encontró nada, devolver última opción razonable
    return 0
# -------------------------
# DETECTOR DE DESALADOR EN UNA COLUMNA
# -------------------------
def buscar_desalador_columna(df, col_idx, filas_adelante=8, filas_detras=8):
    patron = re.compile(r"(c[\-\_ ]?\d{1,3})", flags=re.IGNORECASE)
    nrows = df.shape[0]
    for r in range(0, min(filas_adelante, nrows)):
        try:
            val = df.iloc[r, col_idx]
        except IndexError:
            continue
        if pd.isna(val):
            continue
        s = str(val)
        m = patron.search(s)
        if m:
            return m.group(1).replace("_", "").replace(" ", "").replace("-", "").upper()
    for r in range(0, min(filas_detras, nrows)):
        try:
            val = df.iloc[r, col_idx]
        except IndexError:
            continue
        if pd.isna(val):
            continue
        s = str(val)
        m = patron.search(s)
        if m:
            return m.group(1).replace("_", "").replace(" ", "").replace("-", "").upper()
    for r in range(0, nrows):
        try:
            val = df.iloc[r, col_idx]
        except IndexError:
            continue
        if pd.isna(val):
            continue
        s = str(val)
        m = patron.search(s)
        if m:
            return m.group(1).replace("_", "").replace(" ", "").replace("-", "").upper()
    return ""

# -------------------------
# CONSTRUIR NOMBRES DE COLUMNAS (variable + desalador)
# -------------------------
def construir_nombres_columnas(df_raw, col_inicio=0, col_fin=None, fila_desalador_idx=0, fila_variable_idx=1):
    if col_fin is None:
        col_fin = df_raw.shape[1]
    nombres = []
    desaladores_por_col = []
    for col in range(col_inicio, col_fin):
        desal = df_raw.iloc[fila_desalador_idx, col] if fila_desalador_idx < df_raw.shape[0] else None
        var   = df_raw.iloc[fila_variable_idx, col]  if fila_variable_idx < df_raw.shape[0]  else None

        desal_txt = "" if pd.isna(desal) else str(desal).strip()
        var_txt   = "" if pd.isna(var)   else str(var).strip()

        if desal_txt == "":
            desal_detectado = buscar_desalador_columna(df_raw, col, filas_adelante=6, filas_detras=6)
            if desal_detectado != "":
                desal_txt = desal_detectado

        if var_txt == "":
            for r in range(fila_variable_idx, min(fila_variable_idx + 6, df_raw.shape[0])):
                test = df_raw.iloc[r, col]
                if not pd.isna(test) and str(test).strip() != "":
                    var_txt = str(test).strip()
                    break

        if var_txt == "" and desal_txt == "":
            nombre_final = f"Variable_sin_nombre_{col}"
        elif desal_txt == "":
            nombre_final = f"{var_txt} GENERAL"
        else:
            last_token = desal_txt.strip()
            if var_txt.strip().upper().endswith(last_token.upper()):
                nombre_final = var_txt
            else:
                nombre_final = f"{var_txt} {desal_txt}"

        nombres.append(nombre_final)
        desaladores_por_col.append(desal_txt)

    return nombres, desaladores_por_col

# -------------------------
# MAPEO DE VARIABLES BASE Y NORMALIZACIÓN
# -------------------------
def construir_mapa_variables_base(nombres):
    mapa_variable_a_columnas = {}
    mapa_norm_columns = {}
    for nom in nombres:
        parts = nom.split()
        base = nom
        if len(parts) > 1:
            last = parts[-1]
            if re.match(r"^c[\-]?\d+$", last.strip().lower()) or re.match(r"^c\d+$", last.strip().lower()):
                base = " ".join(parts[:-1]).strip()
        base = base if base != "" else "Variable_sin_nombre"
        mapa_variable_a_columnas.setdefault(base, []).append(nom)
    for base, cols in mapa_variable_a_columnas.items():
        mapa_norm_columns[base] = [(c, clean_token(c)) for c in cols]
    return mapa_variable_a_columnas, mapa_norm_columns

# -------------------------
# LIMPIEZA NUMÉRICA ROBUSTA
# -------------------------
def limpiar_serie_a_numero(serie):
    """
    Limpieza robusta de una serie (pandas Series):
    - Detecta sentinels y mensajes tipo "No Good Data", "No Data", "[12345] No Good Data..." -> NaN
    - Normaliza separadores de miles/decimales
    - Devuelve una Serie numérica (float) con NaN cuando no hay valor válido
    """
    s = serie.astype(str).fillna("").str.strip()

    # Normalizar espacios
    candidato = s.str.replace(r"\s+", " ", regex=True)

    # Patrón sentinel: frases comunes que indican falta de dato
    sentinel_pattern = re.compile(
        r"(no\s+good\s+data|no\s+data|no\s+value|no\s+reading|not\s+available|nodata|n/a|not\s+applicable|no\s+reading)",
        flags=re.IGNORECASE
    )

    # Patrón de código entre corchetes seguido de texto (ej: [-11059] No Good Data For Calculation)
    bracket_code_pattern = re.compile(r"^\s*\[?-?\d+\]?\s*(?:no\b|no good|no data|no value).*", flags=re.IGNORECASE)

    def normaliza_num_str(x):
        if x is None:
            return None
        txt = str(x).strip()
        if txt == "":
            return None

        low = txt.lower()

        # 1) Si coincide con sentinel textual -> NaN
        if sentinel_pattern.search(low) or bracket_code_pattern.match(txt):
            return None

        # 2) Detectar códigos numéricos de sentinel conocidos (ej. -11059, -110)
        if re.fullmatch(r"-?11059|-?110", txt):
            return None

        # 3) Extraer primer número que parezca válido (mantener signos y separadores)
        m = re.search(r"([-+]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d+)?|[-+]?\d*[,\.]?\d+)", txt)
        if not m:
            return None

        numstr = m.group(0)

        # Normalizar separadores: heurística para distinguir miles/decimales
        commas = numstr.count(",")
        dots = numstr.count(".")
        if dots > 0 and commas > 0 and numstr.rfind(",") > numstr.rfind("."):
            # "1.234,56" -> "1234.56"
            s2 = numstr.replace(".", "").replace(",", ".")
            numstr = s2
        elif commas > 0 and dots > 0 and numstr.rfind(".") > numstr.rfind(","):
            # "1,234.56" -> "1234.56"
            numstr = numstr.replace(",", "")
        elif commas > 0 and dots == 0:
            # "1234,56" -> "1234.56"
            numstr = numstr.replace(",", ".")
        # else: "1234.56" ya está bien

        try:
            return float(numstr)
        except:
            return None

    # Aplicar vectorizado
    normalized = candidato.apply(normaliza_num_str)
    numeric = pd.to_numeric(normalized, errors='coerce')

    return numeric

# -------------------------
# LIMPIEZA Y CONSTRUCCIÓN DATAFRAME PRINCIPAL
# -------------------------
def limpiar_dataframe_numerico(datos_base_raw, lista_nombres, df_raw=None, indice_fila_inicio=None, columna_fecha_index=1):
    ncols = datos_base_raw.shape[1]
    if len(lista_nombres) < ncols:
        lista_nombres = lista_nombres + [f"Variable_sin_nombre_extra_{i}" for i in range(ncols - len(lista_nombres))]
    elif len(lista_nombres) > ncols:
        lista_nombres = lista_nombres[:ncols]

    df = datos_base_raw.reset_index(drop=True).copy()
    df.columns = lista_nombres

    for c in df.columns:
        df[c] = limpiar_serie_a_numero(df[c])

    if df_raw is not None and indice_fila_inicio is not None and columna_fecha_index is not None:
        try:
            tiempos = df_raw.iloc[indice_fila_inicio:, columna_fecha_index]
            tiempos = pd.to_datetime(tiempos, errors='coerce')
            tiempos = tiempos.reset_index(drop=True)
            if len(tiempos) < len(df):
                tiempos = tiempos.reindex(range(len(df)))
            elif len(tiempos) > len(df):
                tiempos = tiempos.iloc[:len(df)].reset_index(drop=True)
            if "Tiempo" in df.columns:
                df = df.drop(columns=["Tiempo"])
            df.insert(0, "Tiempo", tiempos)
        except Exception as e:
            print("⚠️ Error al reconstruir Tiempo:", e)

    return df

# -------------------------
# SEPARAR VARIABLES POR DESALADOR
# -------------------------
def separar_variables_por_desalador(columnas, desaladores):
    """
    Agrupa columnas para cada desalador buscando cualquier aparición
    flexible de C11, C-11, 611-C11, ... y variantes.
    """
    grupos = {d: [] for d in desaladores}
    comunes = []

    # Normaliza tokens tipo C11 y crea variantes posibles
    desal_tokens = {}
    for d in desaladores:
        base = normalizar(d)       # C11
        variantes = {
            base,
            base.replace("c", "c-"),   # c-11
            base.replace("c", "611-c"), # 611-c11
            base.replace("c", "c "),    # c 11
            base.replace("-", ""),
        }
        desal_tokens[d] = variantes

    for c in columnas:
        norm = normalizar(c)
        asignado = False

        for d, variantes in desal_tokens.items():
            for v in variantes:
                if v in norm:
                    grupos[d].append(c)
                    asignado = True
                    break
            if asignado:
                break

        if not asignado:
            comunes.append(c)

    return grupos, comunes

# -------------------------
# OBTENER COLUMNA BASE POR DESALADOR (matching)
# -------------------------
def obtener_columnas_base_por_desalador(variable_base, mapa_norm_columns, desaladores):
    resultado = {}
    posibles = mapa_norm_columns.get(variable_base, [])
    if not posibles:
        for d in desaladores:
            resultado[d] = None
        return resultado

    desal_tokens = {d: clean_token(d) for d in desaladores}
    posibles_list = list(posibles)

    for d, dtoken in desal_tokens.items():
        elegido = None
        base_token = clean_token(variable_base)
        for orig, token in posibles_list:
            if token.startswith(base_token) and token.endswith(dtoken):
                elegido = orig
                break
        if elegido is None:
            for orig, token in posibles_list:
                if token.endswith(dtoken) and base_token in token:
                    elegido = orig
                    break
        if elegido is None:
            pattern = base_token + dtoken
            for orig, token in posibles_list:
                if pattern in token:
                    elegido = orig
                    break
        if elegido is None:
            for orig, token in posibles_list:
                if token.endswith(dtoken):
                    elegido = orig
                    break
        if elegido is None:
            for orig, token in posibles_list:
                if dtoken in token:
                    elegido = orig
                    break
        resultado[d] = elegido
        # 🔥 FALLBACK FINAL: si no se encontró coincidencia por desalador,
    # usar la primera columna que coincida con la variable base (caso GENERAL)
    for d in desaladores:
        if resultado[d] is None and len(posibles_list) == 1:
            resultado[d] = posibles_list[0][0]

    return resultado
def calcular_medias_python(df_sub, col_var, base_series, valor_critico):
    serie = pd.to_numeric(df_sub[col_var], errors='coerce')

    mask_sup = base_series > valor_critico
    mask_inf = base_series <= valor_critico

    return {
        "media_total": np.nanmean(serie),
        "std_total":   np.nanstd(serie),
        "media_sup":   np.nanmean(serie[mask_sup]),
        "std_sup":     np.nanstd(serie[mask_sup]),
        "media_inf":   np.nanmean(serie[mask_inf]),
        "std_inf":     np.nanstd(serie[mask_inf]),
        "count_total": np.sum(serie.notna()),
        "count_sup":   np.sum(serie[mask_sup].notna()),
        "count_inf":   np.sum(serie[mask_inf].notna())
    }

# -------------------------
# ANALISIS CRITICO EXTENDIDO (con FORMULAS EXCEL)
# -------------------------
import locale

def analisis_critico_extendido(datos, desaladores, variable_base, valor_critico, carpeta_salida, mapa_norm_columns):

    if 'Tiempo' not in datos.columns:
        raise ValueError("No se encontró la columna 'Tiempo' en los datos.")

    grupos, comunes = separar_variables_por_desalador(list(datos.columns.drop('Tiempo')), desaladores)
    resultados = {}

    for d in desaladores:
        print(f"🔍 Procesando análisis crítico en {d}...")

        cols = grupos.get(d, []) + comunes
        if len(cols) == 0:
            print(f"⚠️ No hay columnas detectadas para {d}. Omitido.")
            continue

        df_sub = datos[['Tiempo'] + cols].copy()

        # Encontrar columna base usada (variable_base)
        posibles = mapa_norm_columns.get(variable_base, [])
        col_desal = None
        d_norm = clean_token(d)

        for original_name, norm_name in posibles:
            if norm_name.endswith(d_norm):
                col_desal = original_name
                break

        if col_desal is None and len(posibles) > 0:
            col_desal = posibles[0][0]

        if col_desal is None:
            print(f"❌ No hay columna base '{variable_base}' para {d}.")
            continue

        for c in cols:
            if c != 'Tiempo':
                df_sub[c] = pd.to_numeric(df_sub[c], errors='coerce')

        base_series = df_sub[col_desal]
        df_arriba = df_sub[base_series > valor_critico].reset_index(drop=True)
        df_abajo  = df_sub[base_series <= valor_critico].reset_index(drop=True)

        # Sanear la variable base para nombre de archivo
        var_base_clean = re.sub(r"[^A-Za-z0-9_-]", "_", variable_base)
        
        archivo = os.path.join(
            carpeta_salida,
            f"Analisis_Critico_{d}_{var_base_clean}.xlsx"
        )
        wb = Workbook()

        # Valores_mayor_crit
        ws_up = wb.active
        ws_up.title = "Valores_mayor_crit"
        for r in dataframe_to_rows(df_arriba, index=False, header=True):
            ws_up.append(r)

        # Valores_menor_igual_crit
        ws_down = wb.create_sheet("Valores_menor_igual_crit")
        for r in dataframe_to_rows(df_abajo, index=False, header=True):
            ws_down.append(r)

        # Todos_los_valores
        ws_all = wb.create_sheet("Todos_los_valores")
        for r in dataframe_to_rows(df_sub, index=False, header=True):
            ws_all.append(r)

        # Resumen Estadístico (sin fórmulas)
        ws_r = wb.create_sheet("Resumen_Estadistico")
        ws_r.append([
            "Variable","Columna_base_usada",
            "Media_total","Std_total",
            "Media_>crit","Std_>crit",
            "Media_<=crit","Std_<=crit",
            "Count_total","Count_>crit","Count_<=crit"
        ])

        for col_var in cols:
            serie = df_sub[col_var]

            vals = {
                "media_total": float(np.nanmean(serie))     if serie.notna().sum() > 0 else None,
                "std_total":   float(np.nanstd(serie))      if serie.notna().sum() > 1 else None,

                "media_sup": float(np.nanmean(serie[base_series > valor_critico]))
                               if (serie[base_series > valor_critico].notna().sum() > 0) else None,

                "std_sup":   float(np.nanstd(serie[base_series > valor_critico]))
                               if (serie[base_series > valor_critico].notna().sum() > 1) else None,

                "media_inf": float(np.nanmean(serie[base_series <= valor_critico]))
                               if (serie[base_series <= valor_critico].notna().sum() > 0) else None,

                "std_inf":   float(np.nanstd(serie[base_series <= valor_critico]))
                               if (serie[base_series <= valor_critico].notna().sum() > 1) else None,

                "count_total": int(serie.notna().sum()),
                "count_sup":   int(serie[base_series > valor_critico].notna().sum()),
                "count_inf":   int(serie[base_series <= valor_critico].notna().sum())
            }

            ws_r.append([
                col_var,
                col_desal,
                vals["media_total"],
                vals["std_total"],
                vals["media_sup"],
                vals["std_sup"],
                vals["media_inf"],
                vals["std_inf"],
                vals["count_total"],
                vals["count_sup"],
                vals["count_inf"]
            ])

        wb.save(archivo)
        print(f"📁 Guardado: {archivo}")

    return resultados

# -------------------------
# GRAFICAS POR DESALADOR (genera .xlsx con imágenes)
# -------------------------
def generar_graficas_por_desalador(datos, desaladores, variable_base, carpeta_salida, mapa_norm_columns):
    grupos, comunes = separar_variables_por_desalador(list(datos.columns.drop('Tiempo')), desaladores)
    mapping_base = obtener_columnas_base_por_desalador(variable_base, mapa_norm_columns, desaladores)

    for d in desaladores:
        print(f"📊 Generando gráficas para {d}...")
        cols = grupos.get(d, []) + comunes
        if len(cols) == 0:
            print(f"⚠️ No hay columnas para {d}. Omito.")
            continue

        col_base = mapping_base.get(d)
        if col_base is None:
            print(f"❌ No se encontró columna base '{variable_base}' para desalador {d}. Omito gráficas.")
            continue

        df_sub = datos[['Tiempo'] + cols].copy()
        for c in cols:
            df_sub[c] = pd.to_numeric(df_sub[c], errors='coerce')
        df_sub[col_base] = pd.to_numeric(df_sub[col_base], errors='coerce')

        wb = Workbook()
        ws0 = wb.active
        ws0.title = "Resumen"
        ws0["A1"] = f"Gráficas desalador {d} (base: {col_base})"

        for c in cols:
            if c == col_base:
                continue
            serie = df_sub[c]
            base = df_sub[col_base]
            tiempo = df_sub["Tiempo"]

            # Solo filtrar valores positivos en la variable base (eficiencia)
            if c == col_base:
                mask = (
                    serie.notna() &
                    base.notna() &
                    tiempo.notna() &
                    (serie.astype(float) > 0)
                )
            else:
                mask = (
                    serie.notna() &
                    base.notna() &
                    tiempo.notna()
                )

            serie_m  = serie[mask]
            base_m   = base[mask]
            tiempo_m = tiempo[mask]

            if len(serie_m) == 0:
                continue

            plt.figure(figsize=(6, 4))
            plt.scatter(base_m, serie_m, s=20, alpha=0.7)
            plt.xlabel(col_base)
            plt.ylabel(c)
            plt.title(f"{c} vs {col_base}")
            plt.grid(True)
            buf1 = io.BytesIO()
            plt.savefig(buf1, format="png", bbox_inches="tight")
            plt.close()
            buf1.seek(0)

            plt.figure(figsize=(6, 4))
            plt.scatter(tiempo_m, serie_m, s=20, alpha=0.7)
            plt.xlabel("Tiempo")
            plt.ylabel(c)
            plt.xticks(rotation=25)
            plt.title(f"{c} vs Tiempo")
            plt.grid(True)
            buf2 = io.BytesIO()
            plt.savefig(buf2, format="png", bbox_inches="tight")
            plt.close()
            buf2.seek(0)

            hoja = re.sub(r"[^A-Za-z0-9_\- ]", "", c)[:31] or "Var"
            ws = wb.create_sheet(title=hoja)
            insertar_imagen_ws(ws, buf1, "A1")
            insertar_imagen_ws(ws, buf2, "I1")

            df_out = pd.DataFrame({"Tiempo": tiempo_m.values, col_base: base_m.values, c: serie_m.values})
            for r, row in enumerate(dataframe_to_rows(df_out, index=False, header=True), start=35):
                for col_i, val in enumerate(row, start=1):
                    ws.cell(row=r, column=col_i, value=val)

        var_base_clean = re.sub(r"[^A-Za-z0-9_-]", "_", variable_base)
        archivo = os.path.join(
            carpeta_salida,
            f"Graficas_{d}_{var_base_clean}.xlsx"
        )
        print(f"Intentando guardar en: {archivo}")
        wb.save(archivo)
        print(f"📁 Guardado: {archivo}")

# -------------------------
# DETECCIÓN SIMPLE DE TOKENS TIPO FECHA
# -------------------------
def es_token_fecha_like(token):
    if token is None:
        return False
    t = str(token)
    if re.match(r"^\d{6,14}$", t):
        return True
    if re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", t):
        return True
    return False

# -------------------------
# FLUJO PRINCIPAL
# -------------------------
def main():
    print("===================================")
    print("🧭 MENÚ PRINCIPAL DE ANÁLISIS")
    print("===================================")
    print("1) Análisis crítico por variable (fórmulas Excel)")
    print("2) Generar gráficas y análisis general")
    op = input("Elige opción (1 o 2): ").strip()

    nombre_hoja = input("Nombre de la hoja (ej: 'Crudo 1 T.def'): ").strip()
    print(f"📘 Leyendo hoja '{nombre_hoja}'...")
    df_raw = leer_hoja_sin_encabezado(RUTA_EXCEL, nombre_hoja)

    col_inicio = 1
    col_fin = df_raw.shape[1]

    fila_inicio = detectar_fila_inicio_datos(df_raw)
    print(f"Datos detectados a partir de fila {fila_inicio+1}")

    fila_desalador_idx = 0
    fila_variable_idx = 1
    if fila_variable_idx >= fila_inicio:
        fila_variable_idx = max(0, fila_inicio - 1)
        fila_desalador_idx = max(0, fila_variable_idx - 1)

    nombres_columnas, desaladores_por_col = construir_nombres_columnas(
        df_raw, col_inicio=col_inicio, col_fin=col_fin,
        fila_desalador_idx=fila_desalador_idx, fila_variable_idx=fila_variable_idx
    )

    print("\n=== COLUMNAS DETECTADAS ===")
    for c in nombres_columnas:
        print(f"{c}  -->  {clean_token(c)}")

    datos_raw_vals = df_raw.iloc[fila_inicio:, col_inicio:col_fin].reset_index(drop=True)

    datos_base = limpiar_dataframe_numerico(datos_raw_vals, nombres_columnas,
                                           df_raw=df_raw, indice_fila_inicio=fila_inicio, columna_fecha_index=1)
    # ======================================================
    # 🔥 ELIMINAR VALORES NEGATIVOS EN "Eficiencia desalación %"
    # ======================================================
    
    # Buscar el nombre exacto de la columna
    col_eff = None
    for c in datos_base.columns:
        if "eficiencia" in c.lower() and "desal" in c.lower():
            col_eff = c
            break
    
    if col_eff:
        # Reemplazar negativos y ceros por NaN
        datos_base[col_eff] = pd.to_numeric(datos_base[col_eff], errors='coerce')
        datos_base.loc[datos_base[col_eff] <= 0, col_eff] = np.nan
        print(f"✔ Columna '{col_eff}': valores negativos/ceros eliminados")
    else:
        print("⚠ No se encontró la columna de eficiencia para limpieza.")

    mapa_variable_a_columnas, mapa_norm_columns = construir_mapa_variables_base(nombres_columnas)
    variables_base = list(mapa_variable_a_columnas.keys())
    # Limpiar variables base: eliminar el sufijo C11, C12, etc.
    variables_base_limpias = []
    for v in variables_base:
        partes = v.split()
        # Si termina en C11, C12, C15... lo quitamos
        if len(partes) > 1 and partes[-1].upper().startswith("C") and partes[-1][1:].isdigit():
            variables_base_limpias.append(" ".join(partes[:-1]))
        else:
            variables_base_limpias.append(v)
    
    # Eliminar duplicados manteniendo orden
    variables_base_filtradas = list(dict.fromkeys(variables_base_limpias))
    variables_base_filtradas = [v for v in variables_base if not es_token_fecha_like(clean_token(v))]

    if len(variables_base_filtradas) == 0:
        print("No se detectaron variables base válidas. Revisa el encabezado del Excel.")
        sys.exit(1)

    print("\nVariables generales detectadas (base):")
    for i, v in enumerate(variables_base_filtradas):
        print(f"{i+1}. {v}")
    try:
        idx = int(input("Elige variable base (número): ").strip()) - 1
        if idx < 0 or idx >= len(variables_base_filtradas):
            raise ValueError
    except:
        print("Selección inválida. Saliendo.")
        sys.exit(1)
    variable_base = variables_base_filtradas[idx]

    hay_varios = input("¿Hay más de un desalador? (s/n): ").strip().lower()
    if hay_varios == "s":
        desaladores_input = input("Introduce nombres (ej: C11,C12) separados por coma: ").strip()
        desaladores = [x.strip().upper() for x in desaladores_input.split(",") if x.strip() != ""]
        if not desaladores:
            desaladores = ["GENERAL"]
    else:
        deducidos = set()
        for n in nombres_columnas:
            parts = n.split()
            if len(parts) > 0:
                last = parts[-1]
                if re.match(r"^c[\-]?\d+", last.strip().lower()):
                    ded = re.sub(r"[\-_\s]", "", last).upper()
                    deducidos.add(ded)
        if len(deducidos) > 0:
            # FORZAR que 1 solo desalador = GENERAL
            if len(deducidos) == 1:
                desaladores = ["GENERAL"]
            else:
                desaladores = sorted(list(deducidos))

            print("Se han detectado desaladores:", desaladores)
        else:
            desaladores = ["GENERAL"]

    print(f"Desaladores: {desaladores}")

    if op == "1":
        try:
            valor_critico = float(input(f"Introduce valor crítico para '{variable_base}': ").replace(",", "."))
        except:
            print("Valor crítico inválido.")
            sys.exit(1)
        carpeta = os.path.join(DIRECTORIO_SALIDA, "Analisis_Criticos")
        os.makedirs(carpeta, exist_ok=True)
        analisis_critico_extendido(datos_base, desaladores, variable_base, valor_critico, carpeta, mapa_norm_columns)
        print("🎉 Análisis crítico completado.")
        sys.exit(0)

    elif op == "2":
        carpeta = os.path.join(DIRECTORIO_SALIDA, "Graficas")
        os.makedirs(carpeta, exist_ok=True)
        generar_graficas_por_desalador(datos_base, desaladores, variable_base, carpeta, mapa_norm_columns)
        print("🎉 Gráficas completadas.")
        sys.exit(0)

    else:
        print("Opción inválida.")
        sys.exit(1)

if __name__ == "__main__":
    main()
