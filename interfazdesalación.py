# app_desalacion_unificado_largo_tabs.py
# ---------------------------------------
# Versión reorganizada con dos pestañas:
#   1) Graficado → contiene TODO el código original
#   2) Análisis  → Random Forest + resumen + gráfico de importancias
# ---------------------------------------

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.ensemble import RandomForestRegressor
# para nuevos modelos y explicadores
import shap
shap.initjs()  # no hace nada en Streamlit pero evita warnings
from lime.lime_tabular import LimeTabularExplainer
import xgboost as xgb
import lightgbm as lgb
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.linear_model import Ridge, Lasso
import streamlit.components.v1 as components
import io

# =============================================================
# CONFIGURACIÓN BÁSICA
# =============================================================
st.set_page_config(page_title="Análisis desaladoras", layout="wide")

st.markdown("<h1 class='darkblue-title'>Análisis desaladoras</h1>", unsafe_allow_html=True)

# =============================================================
# PESTAÑAS PRINCIPALES
# =============================================================
tab_graf, tab_rf = st.tabs(["Graficado", "Análisis"])

# =============================================================
# 1) PESTAÑA GRAFICADO (TODO EL CÓDIGO ORIGINAL)
# =============================================================
with tab_graf:
    # app_desalacion_unificado_largo.py
    # Interfaz Streamlit unificada y extensa para anÃ¡lisis de desaladoras
    # - Integra la lÃ³gica del 'Programa Eficiencias de desalacion2.py'
    # - Reproduce exactamente la construcciÃ³n de "variables base" que usa el programa principal
    # - Lee Excel sin cabeceras fijas y procesa todos los datos
    # - Si hay +1 desalador, el desplegable mostrarÃ¡ solo los nombres bÃ¡sicos (sin C11, etc.)
    #
    # Guarda como app_desalacion_unificado_largo.py y ejecuta:
    #    streamlit run app_desalacion_unificado_largo.py
    #
    # Autor: IntegraciÃ³n a partir de los archivos proporcionados por el usuario.
    # Fecha: generada automÃ¡ticamente por el asistente.

    import os
    import sys
    import re
    import io
    import unicodedata
    import tempfile
    import importlib.util
    from pathlib import Path
    from typing import List, Tuple, Dict, Any

    import numpy as np
    import pandas as pd
    import matplotlib.pyplot as plt
    from PIL import Image, ImageFilter

    import streamlit as st
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.image import Image as xlImage

    # -------------------------
    # ConfiguraciÃ³n de la app
    # -------------------------
    # -------------------------
    # ConfiguraciÃ³n bÃ¡sica y estilo
    # -------------------------


    # Estilo global: colores, headers, botones
    st.markdown("""
    <style>

      /* =========================================================
         0. FONDO GENERAL â BLANCO
         ========================================================= */
      html, body, .block-container, [class*="stApp"] {
          background-color: #FFFFFF !important;  /* blanco */
          color: #333333 !important;             /* texto gris oscuro */
      }

      /* =========================================================
         1. TITULOS GRANDES â NARANJA REPSOL
         ========================================================= */
      h1, h2, h3, h4, h5, h6 {
          color: #D98B3B !important;     /* naranja Repsol */
          font-weight: 800 !important;
      }

      /* =========================================================
         2. TITULOS AZUL OSCURO (solo si tÃº lo marcas con clase)
         ========================================================= */
      .darkblue-title {
          color: #0B1A33 !important;     /* azul oscuro */
          font-weight: 800 !important;
      }

      /* =========================================================
         3. WIDGETS â letra gris oscuro
         ========================================================= */
      .stSelectbox label,
      .stMultiSelect label,
      .stNumberInput label,
      .stSlider label,
      .stTextInput label {
          color: #333333 !important;
      }

      .css-16idsys, .css-1pndypt, .css-1offfwp, .css-1kyxreq {
          color: #333333 !important;
      }

      .stSelectbox div[data-baseweb="select"],
      .stMultiSelect div[data-baseweb="select"] {
          background-color: white !important;
          color: #333333 !important;
      }

      /* =========================================================
         4. TABS â gris / ROJO seleccionada
         ========================================================= */
      .stTabs [data-baseweb="tab"] p {
          color: #666666 !important;   /* gris */
          font-weight: 600 !important;
      }

      .stTabs [aria-selected="true"] p {
          color: red !important;       /* ROJO al seleccionar */
          font-weight: 700 !important;
      }

      .stTabs [data-baseweb="tab"] {
          background-color: #FFFFFF !important; /* fondo blanco */
      }

      /* =========================================================
         5. Botones â NARANJAS
         ========================================================= */
      .stButton>button {
          background-color: #D98B3B !important;
          color: white !important;
          border-radius: 8px;
      }
      .stButton>button:hover {
          background-color: #b57830 !important;
          color: white !important;
      }

    </style>
    """, unsafe_allow_html=True)

    # -------------------------
    # Intento cargar mÃ³dulo original (si existe en /mnt/data)
    # -------------------------
    MODULE_PATH = Path("/mnt/data/Programa Eficiencias de desalacion2.py")
    user_mod = None
    if MODULE_PATH.exists():
        try:
            spec = importlib.util.spec_from_file_location("prog_desal", str(MODULE_PATH))
            user_mod = importlib.util.module_from_spec(spec)
            sys.modules["prog_desal"] = user_mod
            spec.loader.exec_module(user_mod)
            st.sidebar.success(f"Módulo original cargado desde {MODULE_PATH}")
        except Exception as e:
            st.sidebar.error(f"No se pudo cargar módulo original: {e}")
    else:
        st.sidebar.info("No se encontró el módulo original en /mnt/data; utilizando implementaciones internas.")

    def safe_get(name, fallback=None):
        """Si se carga el módulo original, devuelve la función exportada; si no, devuelve fallback."""
        if user_mod is None:
            return fallback
        return getattr(user_mod, name, fallback)

    # -------------------------
    # Utilidades (tomadas de tu cÃ³digo original)
    # -------------------------
    def normalizar(txt):
        if txt is None:
            return ""
        txt = str(txt).strip().lower()
        txt = ''.join(c for c in unicodedata.normalize('NFD', txt) if unicodedata.category(c) != 'Mn')
        return txt

    def clean_token(s):
        if s is None:
            return ""
        s = str(s)
        s = ''.join(ch for ch in unicodedata.normalize('NFD', s) if unicodedata.category(ch) != 'Mn')
        s = re.sub(r"[^A-Za-z0-9]", "", s)
        return s.lower()

    def make_unique(col_list: List[str]) -> List[str]:
        """Convierte nombres a Ãºnicos añadiendo sufijos __N cuando sea necesario"""
        seen = {}
        out = []
        for c in col_list:
            key = str(c)
            if key not in seen:
                seen[key] = 1
                out.append(key)
            else:
                seen[key] += 1
                out.append(f"{key}__{seen[key]}")
        return out

    def insertar_imagen_ws(ws, buf, posicion="A1"):
        try:
            img = xlImage(buf)
            img.anchor = posicion
            ws.add_image(img)
        except Exception as e:
            print("Error insertando imagen en worksheet:", e)

    # -------------------------
    # Funciones para leer y detectar
    # -------------------------
    def leer_hoja_sin_encabezado(path_excel: str, nombre_hoja: str) -> pd.DataFrame:
        """Lee hoja sin encabezados (header=None) usando openpyxl"""
        df_raw = pd.read_excel(path_excel, sheet_name=nombre_hoja, header=None, engine="openpyxl")
        return df_raw

    def detectar_fila_inicio_datos_fallback(df_raw: pd.DataFrame) -> int:
        """
        Heurí­stica para detectar la fila donde comienzan los datos.
        Copiado de tu programa original.
        """
        palabras_ruido = [
            "media", "desviacion", "max", "min",
            "servidor", "unidades", "escala", "ph",
            "tension", "consumo", "eficiencia"
        ]

        nfilas, ncols = df_raw.shape
        for i in range(nfilas):
            fila = df_raw.iloc[i, :]
            texto_fila = " ".join(str(x).lower() for x in fila if pd.notna(x))
            if any(p in texto_fila for p in palabras_ruido):
                continue

            num_ok = 0
            date_ok = 0

            for v in fila:
                if pd.isna(v):
                    continue
                try:
                    float(str(v).replace(",", "."))
                    num_ok += 1
                    continue
                except:
                    pass
                if isinstance(v, (pd.Timestamp,)):
                    date_ok += 1
                    continue
                try:
                    pd.to_datetime(v, errors="raise")
                    date_ok += 1
                except:
                    pass

            col1 = fila.iloc[1] if len(fila) > 1 else None
            col1_es_fecha = False
            try:
                pd.to_datetime(col1, errors="raise")
                col1_es_fecha = True
            except:
                pass

            if col1_es_fecha:
                return i
            if num_ok >= max(3, int(ncols * 0.40)):
                return i

        return 0

    # Puede venir del mÃ³dulo original
    detectar_fila_inicio_datos = safe_get('detectar_fila_inicio_datos', detectar_fila_inicio_datos_fallback)

    # -------------------------
    # Detector desalador en columna
    # -------------------------
    def buscar_desalador_columna_fallback(df, col_idx, filas_adelante=8, filas_detras=8):
        patron = re.compile(r"(c[\-\_ ]?\d{1,3})", flags=re.IGNORECASE)
        nrows = df.shape[0]
        for r in range(0, min(filas_adelante, nrows)):
            try:
                val = df.iloc[r, col_idx]
            except IndexError:
                continue
            if pd.isna(val):
                continue
            s = str(val)
            m = patron.search(s)
            if m:
                return m.group(1).replace("_", "").replace(" ", "").replace("-", "").upper()
        for r in range(0, min(filas_detras, nrows)):
            try:
                val = df.iloc[r, col_idx]
            except IndexError:
                continue
            if pd.isna(val):
                continue
            s = str(val)
            m = patron.search(s)
            if m:
                return m.group(1).replace("_", "").replace(" ", "").replace("-", "").upper()
        for r in range(0, nrows):
            try:
                val = df.iloc[r, col_idx]
            except IndexError:
                continue
            if pd.isna(val):
                continue
            s = str(val)
            m = patron.search(s)
            if m:
                return m.group(1).replace("_", "").replace(" ", "").replace("-", "").upper()
        return ""

    buscar_desalador_columna = safe_get('buscar_desalador_columna', buscar_desalador_columna_fallback)

    # -------------------------
    # Construir nombres columnas (variable + desalador)
    # -------------------------
    def construir_nombres_columnas_fallback(df_raw, col_inicio=0, col_fin=None, fila_desalador_idx=0, fila_variable_idx=1):
        if col_fin is None:
            col_fin = df_raw.shape[1]
        nombres = []
        desaladores_por_col = []
        for col in range(col_inicio, col_fin):
            desal = df_raw.iloc[fila_desalador_idx, col] if fila_desalador_idx < df_raw.shape[0] else None
            var   = df_raw.iloc[fila_variable_idx, col]  if fila_variable_idx < df_raw.shape[0]  else None

            desal_txt = "" if pd.isna(desal) else str(desal).strip()
            var_txt   = "" if pd.isna(var)   else str(var).strip()

            if desal_txt == "":
                desal_detectado = buscar_desalador_columna(df_raw, col, filas_adelante=6, filas_detras=6)
                if desal_detectado != "":
                    desal_txt = desal_detectado

            if var_txt == "":
                for r in range(fila_variable_idx, min(fila_variable_idx + 6, df_raw.shape[0])):
                    test = df_raw.iloc[r, col]
                    if not pd.isna(test) and str(test).strip() != "":
                        var_txt = str(test).strip()
                        break

            if var_txt == "" and desal_txt == "":
                nombre_final = f"Variable_sin_nombre_{col}"
            elif desal_txt == "":
                nombre_final = f"{var_txt} GENERAL"
            else:
                last_token = desal_txt.strip()
                if var_txt.strip().upper().endswith(last_token.upper()):
                    nombre_final = var_txt
                else:
                    nombre_final = f"{var_txt} {desal_txt}"

            nombres.append(nombre_final)
            desaladores_por_col.append(desal_txt)

        return nombres, desaladores_por_col

    construir_nombres_columnas = safe_get('construir_nombres_columnas', construir_nombres_columnas_fallback)

    # -------------------------
    # Mapeo variables base y normalizaciÃ³n
    # -------------------------
    def construir_mapa_variables_base_fallback(nombres: List[str]) -> Tuple[Dict[str, List[str]], Dict[str, List[Tuple[str,str]]]]:
        mapa_variable_a_columnas = {}
        mapa_norm_columns = {}
        for nom in nombres:
            parts = str(nom).split()
            base = nom
            if len(parts) > 1:
                last = parts[-1]
                if re.match(r"^c[\-]?\d+$", last.strip().lower()) or re.match(r"^c\d+$", last.strip().lower()):
                    base = " ".join(parts[:-1]).strip()
            base = base if base != "" else "Variable_sin_nombre"
            mapa_variable_a_columnas.setdefault(base, []).append(nom)
        for base, cols in mapa_variable_a_columnas.items():
            mapa_norm_columns[base] = [(c, clean_token(c)) for c in cols]
        return mapa_variable_a_columnas, mapa_norm_columns

    construir_mapa_variables_base = safe_get('construir_mapa_variables_base', construir_mapa_variables_base_fallback)

    # -------------------------
    # Limpieza numÃ©rica robusta
    # -------------------------
    def limpiar_serie_a_numero_fallback(serie: pd.Series) -> pd.Series:
        s = serie.astype(str).fillna("").str.strip()
        candidato = s.str.replace(r"\s+", " ", regex=True)
        sentinel_pattern = re.compile(
            r"(no\s+good\s+data|no\s+data|no\s+value|no\s+reading|not\s+available|nodata|n/a|not\s+applicable)",
            flags=re.IGNORECASE
        )
        bracket_code_pattern = re.compile(r"^\s*\[?-?\d+\]?\s*(?:no\b|no good|no data|no value).*", flags=re.IGNORECASE)

        def normaliza_num_str(x):
            if x is None:
                return None
            txt = str(x).strip()
            if txt == "":
                return None
            low = txt.lower()
            if sentinel_pattern.search(low) or bracket_code_pattern.match(txt):
                return None
            if re.fullmatch(r"-?11059|-?110", txt):
                return None
            m = re.search(r"([-+]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d+)?|[-+]?\d*[,\.]?\d+)", txt)
            if not m:
                return None
            numstr = m.group(0)
            commas = numstr.count(",")
            dots = numstr.count(".")
            if dots > 0 and commas > 0 and numstr.rfind(",") > numstr.rfind("."):
                s2 = numstr.replace(".", "").replace(",", ".")
                numstr = s2
            elif commas > 0 and dots > 0 and numstr.rfind(".") > numstr.rfind(","):
                numstr = numstr.replace(",", "")
            elif commas > 0 and dots == 0:
                numstr = numstr.replace(",", ".")
            try:
                return float(numstr)
            except:
                return None

        normalized = candidato.apply(normaliza_num_str)
        numeric = pd.to_numeric(normalized, errors='coerce')
        return numeric

    limpiar_serie_a_numero = safe_get('limpiar_serie_a_numero', limpiar_serie_a_numero_fallback)

    # -------------------------
    # Limpieza y construcciÃ³n DataFrame principal
    # -------------------------
    def limpiar_dataframe_numerico_fallback(datos_base_raw: pd.DataFrame, lista_nombres: List[str],
                                           df_raw: pd.DataFrame=None, indice_fila_inicio: int=None,
                                           columna_fecha_index: int=1) -> pd.DataFrame:
        ncols = datos_base_raw.shape[1]
        if len(lista_nombres) < ncols:
            lista_nombres = lista_nombres + [f"Variable_sin_nombre_extra_{i}" for i in range(ncols - len(lista_nombres))]
        elif len(lista_nombres) > ncols:
            lista_nombres = lista_nombres[:ncols]

        df = datos_base_raw.reset_index(drop=True).copy()
        df.columns = lista_nombres

        for c in df.columns:
            df[c] = limpiar_serie_a_numero(df[c])

        if df_raw is not None and indice_fila_inicio is not None and columna_fecha_index is not None:
            try:
                tiempos = df_raw.iloc[indice_fila_inicio:, columna_fecha_index]
                tiempos = pd.to_datetime(tiempos, errors='coerce')
                tiempos = tiempos.reset_index(drop=True)
                if len(tiempos) < len(df):
                    tiempos = tiempos.reindex(range(len(df)))
                elif len(tiempos) > len(df):
                    tiempos = tiempos.iloc[:len(df)].reset_index(drop=True)
                if "Tiempo" in df.columns:
                    df = df.drop(columns=["Tiempo"])
                df.insert(0, "Tiempo", tiempos)
            except Exception as e:
                print("â ï¸ Error al reconstruir Tiempo:", e)
        return df

    limpiar_dataframe_numerico = safe_get('limpiar_dataframe_numerico', limpiar_dataframe_numerico_fallback)

    # -------------------------
    # Separar variables por desalador
    # -------------------------
    def separar_variables_por_desalador(columnas: List[str], desaladores: List[str]) -> Tuple[Dict[str,List[str]], List[str]]:
        grupos = {d: [] for d in desaladores}
        comunes = []
        desal_tokens = {}
        for d in desaladores:
            base = normalizar(d)
            variantes = {
                base,
                base.replace("c", "c-"),
                base.replace("c", "611-c"),
                base.replace("c", "c "),
                base.replace("-", "")
            }
            desal_tokens[d] = variantes
        for c in columnas:
            norm = normalizar(c)
            asignado = False
            for d, variantes in desal_tokens.items():
                for v in variantes:
                    if v in norm:
                        grupos[d].append(c)
                        asignado = True
                        break
                if asignado:
                    break
            if not asignado:
                comunes.append(c)
        return grupos, comunes

    # -------------------------
    # Obtener columna base por desalador (matching)
    # -------------------------
    def obtener_columnas_base_por_desalador(variable_base: str, mapa_norm_columns: Dict[str, List[Tuple[str,str]]],
                                            desaladores: List[str]) -> Dict[str, Any]:
        resultado = {}
        posibles = mapa_norm_columns.get(variable_base, [])
        if not posibles:
            for d in desaladores:
                resultado[d] = None
            return resultado
        desal_tokens = {d: clean_token(d) for d in desaladores}
        posibles_list = list(posibles)
        for d, dtoken in desal_tokens.items():
            elegido = None
            base_token = clean_token(variable_base)
            for orig, token in posibles_list:
                if token.startswith(base_token) and token.endswith(dtoken):
                    elegido = orig
                    break
            if elegido is None:
                for orig, token in posibles_list:
                    if token.endswith(dtoken) and base_token in token:
                        elegido = orig
                        break
            if elegido is None:
                pattern = base_token + dtoken
                for orig, token in posibles_list:
                    if pattern in token:
                        elegido = orig
                        break
            if elegido is None:
                for orig, token in posibles_list:
                    if token.endswith(dtoken):
                        elegido = orig
                        break
            if elegido is None:
                for orig, token in posibles_list:
                    if dtoken in token:
                        elegido = orig
                        break
            resultado[d] = elegido
        for d in desaladores:
            if resultado[d] is None and len(posibles_list) == 1:
                resultado[d] = posibles_list[0][0]
        return resultado

    # -------------------------
    # AnÃ¡lisis crÃ­tico extendido (internal)
    # -------------------------
    def analisis_critico_extendido_internal(datos: pd.DataFrame, desaladores: List[str], variable_base: str,
                                            valor_critico: float, carpeta_salida: str, mapa_norm_columns: Dict[str, List[Tuple[str,str]]]):
        if 'Tiempo' not in datos.columns:
            raise ValueError("No se encontró la columna 'Tiempo' en los datos.")

        grupos, comunes = separar_variables_por_desalador(list(datos.columns.drop('Tiempo')), desaladores)
        resultados = {}
        os.makedirs(carpeta_salida, exist_ok=True)

        for d in desaladores:
            cols = grupos.get(d, []) + comunes
            if len(cols) == 0:
                print(f"No hay columnas detectadas para {d}.")
                continue
            df_sub = datos[['Tiempo'] + cols].copy()
            posibles = mapa_norm_columns.get(variable_base, [])
            col_desal = None
            d_norm = clean_token(d)
            for original_name, norm_name in posibles:
                if norm_name.endswith(d_norm):
                    col_desal = original_name
                    break
            if col_desal is None and len(posibles) > 0:
                col_desal = posibles[0][0]
            if col_desal is None:
                print(f"No hay columna base '{variable_base}' para {d}.")
                continue
            for c in cols:
                if c != 'Tiempo':
                    df_sub[c] = pd.to_numeric(df_sub[c], errors='coerce')
            base_series = df_sub[col_desal]
            df_arriba = df_sub[base_series > valor_critico].reset_index(drop=True)
            df_abajo  = df_sub[base_series <= valor_critico].reset_index(drop=True)
            var_base_clean = re.sub(r"[^A-Za-z0-9_-]", "_", variable_base)
            archivo = os.path.join(carpeta_salida, f"Analisis_Critico_{d}_{var_base_clean}.xlsx")
            wb = Workbook()
            ws_up = wb.active
            ws_up.title = "Valores_mayor_crit"
            for r in dataframe_to_rows(df_arriba, index=False, header=True):
                ws_up.append(r)
            ws_down = wb.create_sheet("Valores_menor_igual_crit")
            for r in dataframe_to_rows(df_abajo, index=False, header=True):
                ws_down.append(r)
            ws_all = wb.create_sheet("Todos_los_valores")
            for r in dataframe_to_rows(df_sub, index=False, header=True):
                ws_all.append(r)
            ws_r = wb.create_sheet("Resumen_Estadistico")
            ws_r.append([
                "Variable","Columna_base_usada",
                "Media_total","Std_total",
                "Media_>crit","Std_>crit",
                "Media_<=crit","Std_<=crit",
                "Count_total","Count_>crit","Count_<=crit"
            ])
            for col_var in cols:
                serie = df_sub[col_var]
                vals = {
                    "media_total": float(np.nanmean(serie))     if serie.notna().sum() > 0 else None,
                    "std_total":   float(np.nanstd(serie))      if serie.notna().sum() > 1 else None,
                    "media_sup": float(np.nanmean(serie[base_series > valor_critico])) if (serie[base_series > valor_critico].notna().sum() > 0) else None,
                    "std_sup":   float(np.nanstd(serie[base_series > valor_critico])) if (serie[base_series > valor_critico].notna().sum() > 1) else None,
                    "media_inf": float(np.nanmean(serie[base_series <= valor_critico])) if (serie[base_series <= valor_critico].notna().sum() > 0) else None,
                    "std_inf":   float(np.nanstd(serie[base_series <= valor_critico])) if (serie[base_series <= valor_critico].notna().sum() > 1) else None,
                    "count_total": int(serie.notna().sum()),
                    "count_sup":   int(serie[base_series > valor_critico].notna().sum()),
                    "count_inf":   int(serie[base_series <= valor_critico].notna().sum())
                }
                ws_r.append([
                    col_var,
                    col_desal,
                    vals["media_total"],
                    vals["std_total"],
                    vals["media_sup"],
                    vals["std_sup"],
                    vals["media_inf"],
                    vals["std_inf"],
                    vals["count_total"],
                    vals["count_sup"],
                    vals["count_inf"]
                ])
            wb.save(archivo)
            resultados[d] = archivo
        return resultados

    analisis_critico_extendido = safe_get('analisis_critico_extendido', analisis_critico_extendido_internal)

    # -------------------------
    # Generar graficas y guardar .xlsx con imagenes
    # -------------------------
    def generar_graficas_por_desalador_internal(datos: pd.DataFrame, desaladores: List[str], variable_base: str,
                                                carpeta_salida: str, mapa_norm_columns: Dict[str, List[Tuple[str,str]]]):
        grupos, comunes = separar_variables_por_desalador(list(datos.columns.drop('Tiempo')), desaladores)
        mapping_base = obtener_columnas_base_por_desalador(variable_base, mapa_norm_columns, desaladores)
        os.makedirs(carpeta_salida, exist_ok=True)
        resultados = {}
        for d in desaladores:
            cols = grupos.get(d, []) + comunes
            if len(cols) == 0:
                print(f"No hay columnas para {d}.")
                continue
            col_base = mapping_base.get(d)
            if col_base is None:
                print(f"No se encontró columna base '{variable_base}' para desalador {d}.")
                continue
            df_sub = datos[['Tiempo'] + cols].copy()
            for c in cols:
                df_sub[c] = pd.to_numeric(df_sub[c], errors='coerce')
            df_sub[col_base] = pd.to_numeric(df_sub[col_base], errors='coerce')
            wb = Workbook()
            ws0 = wb.active
            ws0.title = "Resumen"
            ws0["A1"] = f"Gráficas desalador {d} (base: {col_base})"
            for c in cols:
                if c == col_base:
                    continue
                serie = df_sub[c]
                base = df_sub[col_base]
                tiempo = df_sub["Tiempo"]
                mask = (serie.notna() & base.notna() & tiempo.notna())
                serie_m  = serie[mask]
                base_m   = base[mask]
                tiempo_m = tiempo[mask]
                if len(serie_m) == 0:
                    continue
                plt.figure(figsize=(6,4))
                plt.scatter(base_m, serie_m, s=20, alpha=0.7)
                plt.xlabel(col_base)
                plt.ylabel(c)
                plt.title(f"{c} vs {col_base}")
                plt.grid(True)
                buf1 = io.BytesIO()
                plt.savefig(buf1, format="png", bbox_inches="tight")
                plt.close()
                buf1.seek(0)
                plt.figure(figsize=(6,4))
                plt.scatter(tiempo_m, serie_m, s=20, alpha=0.7)
                plt.xlabel("Tiempo")
                plt.ylabel(c)
                plt.xticks(rotation=25)
                plt.title(f"{c} vs Tiempo")
                plt.grid(True)
                buf2 = io.BytesIO()
                plt.savefig(buf2, format="png", bbox_inches="tight")
                plt.close()
                buf2.seek(0)
                hoja = re.sub(r"[^A-Za-z0-9_\- ]", "", c)[:31] or "Var"
                ws = wb.create_sheet(title=hoja)
                insertar_imagen_ws(ws, buf1, "A1")
                insertar_imagen_ws(ws, buf2, "I1")
                df_out = pd.DataFrame({"Tiempo": tiempo_m.values, col_base: base_m.values, c: serie_m.values})
                for r, row in enumerate(dataframe_to_rows(df_out, index=False, header=True), start=35):
                    for col_i, val in enumerate(row, start=1):
                        ws.cell(row=r, column=col_i, value=val)
            var_base_clean = re.sub(r"[^A-Za-z0-9_-]", "_", variable_base)
            archivo = os.path.join(carpeta_salida, f"Graficas_{d}_{var_base_clean}.xlsx")
            wb.save(archivo)
            resultados[d] = archivo
        return resultados

    generar_graficas_por_desalador = safe_get('generar_graficas_por_desalador', generar_graficas_por_desalador_internal)

    # -------------------------
    # DetecciÃ³n simple tokens tipo fecha
    # -------------------------
    def es_token_fecha_like(token):
        if token is None:
            return False
        t = str(token)
        if re.match(r"^\d{6,14}$", t):
            return True
        if re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", t):
            return True
        # otras heurÃ­sticas:
        if re.match(r"^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$", t):
            return True
        return False

    # -------------------------
    # UI: Sidebar
    # -------------------------
    st.sidebar.header("Entradas")
    uploaded = st.sidebar.file_uploader("Sube archivo Excel de desalación", type=["xlsx", "xls"], help="Archivo con la estructura del programa original (se leen todas las filas)")
    st.sidebar.markdown("---")
    st.sidebar.header("Parámetros visuales")
    fig_w = st.sidebar.slider("Ancho figura", 6, 18, 10)
    fig_h = st.sidebar.slider("Alto figura", 4, 12, 6)
    st.sidebar.markdown("---")
    # Mostrar logo opcional si estÃ¡
    logo_path = Path("logo_repsol.png")
    if logo_path.exists():
        try:
            logo_original = Image.open(logo_path).convert("RGBA")
            blur_radius = 8
            padding = blur_radius * 3
            new_size = (logo_original.width + padding, logo_original.height + padding)
            final_logo = Image.new("RGBA", new_size, (255,255,255,0))
            center_x = (new_size[0] - logo_original.width) // 2
            center_y = (new_size[1] - logo_original.height) // 2
            final_logo.paste(logo_original, (center_x, center_y), logo_original)
            mask = final_logo.split()[3]
            white_halo = Image.new("RGBA", final_logo.size, (255, 255, 255, 0))
            white_halo.putalpha(mask.filter(ImageFilter.GaussianBlur(blur_radius)))
            final_logo = Image.alpha_composite(white_halo, final_logo)
            st.image(final_logo, width=140)
        except Exception:
            st.info("Error cargando logo_repsol.png")

    # -------------------------
    # Main: cuando hay upload
    # -------------------------
    if uploaded is None:
        st.info("Sube un archivo Excel para comenzar. La app leerá todas las filas y reconstruirá nombres y variables.")
    else:
        # Guardar temporalmente
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpf:
            tmpf.write(uploaded.getbuffer())
            tmp_path = tmpf.name

        try:
            xls = pd.ExcelFile(tmp_path, engine="openpyxl")
            hojas = xls.sheet_names
        except Exception as e:
            st.error(f"Error leyendo Excel: {e}")
            hojas = []

        if hojas:
            hoja_sel = st.selectbox("Selecciona hoja", hojas)
            try:
                df_raw = pd.read_excel(tmp_path, sheet_name=hoja_sel, header=None, engine="openpyxl")
                st.success(f"Hoja '{hoja_sel}' leí­da: filas={df_raw.shape[0]} columnas={df_raw.shape[1]}")
            except Exception as e:
                st.error(f"Error leyendo hoja seleccionada: {e}")
                df_raw = None

            if df_raw is not None:
                # Detectar fila inicio usando la funciÃ³n real si estÃ¡ (o fallback)
                try:
                    fila_inicio = detectar_fila_inicio_datos(df_raw)
                except Exception:
                    fila_inicio = detectar_fila_inicio_datos_fallback(df_raw)
                st.write(f"Fila de inicio detectada (index base 0): {fila_inicio}")

                # Determinar Ã­ndices de filas donde podrÃ­amos tener desalador/variable
                # ============================================
                # ğ§ BLOQUE CORREGIDO PARA FIJAR ENCABEZADOS
                # ============================================
                
                fila_desalador_idx = 0
                fila_variable_idx = 1
                fila_inicio = detectar_fila_inicio_datos(df_raw)

                # Construir nombres de columnas (usando funciÃ³n original si existe)
                try:
                    nombres_col, desaladores_por_col = construir_nombres_columnas(df_raw, col_inicio=1, col_fin=df_raw.shape[1],
                                                                                  fila_desalador_idx=fila_desalador_idx,
                                                                                  fila_variable_idx=fila_variable_idx)
                except Exception:
                    nombres_col, desaladores_por_col = construir_nombres_columnas_fallback(df_raw, col_inicio=1, col_fin=df_raw.shape[1],
                                                                                           fila_desalador_idx=fila_desalador_idx,
                                                                                           fila_variable_idx=fila_variable_idx)

                # Extraer datos desde fila_inicio (todas las filas)
                datos_vals = df_raw.iloc[fila_inicio:, 1:df_raw.shape[1]].reset_index(drop=True)
                for c in datos_vals.columns:
                    datos_vals[c] = limpiar_serie_a_numero(datos_vals[c])
                datos = datos_vals.copy()
                # ======================================================
                # ğ¥ FILTRO STREAMLIT-SEGURO: SOLO EFICIENCIA POSITIVA
                # ======================================================
                
                # Localizar columna de eficiencia (busca cualquier nombre que contenga estos tokens)
                col_eff = None
                
                for c in datos.columns:
                    # asegurar que c es texto
                    cl = str(c).lower()
                
                    if "eficiencia" in cl and "desal" in cl:
                        col_eff = c
                        break
                
                if col_eff:
                    datos[col_eff] = pd.to_numeric(datos[col_eff], errors='coerce')
                    datos.loc[datos[col_eff] <= 0, col_eff] = np.nan
                    st.info(f"Filtro aplicado: valores <= 0 eliminados en '{col_eff}'")

                
                if col_eff:
                    # Convertir a numÃ©rico
                    datos[col_eff] = pd.to_numeric(datos[col_eff], errors='coerce')
                    # Aplicar filtro SOLO en la eficiencia
                    datos.loc[datos[col_eff] <= 0, col_eff] = np.nan
                    st.info(f"Filtro aplicado: valores <= 0 eliminados en '{col_eff}'")

                # asegurar unicidad columnas
                datos.columns = make_unique(nombres_col[:datos.shape[1]])
                # reconstruir columna Tiempo desde la columna 1 (index 1) del raw
                tiempo_col = pd.to_datetime(df_raw.iloc[fila_inicio:, 1].reset_index(drop=True), errors='coerce')
                if tiempo_col.isnull().all():
                    datos.insert(0, "Tiempo", pd.RangeIndex(start=0, stop=len(datos)))
                else:
                    datos.insert(0, "Tiempo", tiempo_col)

                st.subheader("Datos (vista previa)")
                st.dataframe(datos.head(200))
                st.session_state["datos"] = datos

                # Construir mapa variable -> columnas
                try:
                    mapa_variable_a_columnas, mapa_norm_columns = construir_mapa_variables_base(nombres_col)
                except Exception:
                    mapa_variable_a_columnas, mapa_norm_columns = construir_mapa_variables_base_fallback(nombres_col)

                # Detectar desaladores presentes a partir de nombres_col (patrÃ³n C#)
                # ======================================================
                # ğ· Pregunta al usuario si quiere buscar varios desaladores
                # ======================================================
                
                st.sidebar.markdown("---")
                forzar_general = st.sidebar.radio(
                    "¿Quieres que la app busque varios desaladores?",
                    ["No, usar un solo desalador (GENERAL)", "Detectar varios desaladores automáticamente"],
                    index=0
                )

                # ======================================================
                # ğ¥ DETECCIÃN AUTOMÃTICA DE DESALADORES (STREAMLIT)
                # ======================================================
                
                desaladores_detectados = set()
                
                patron_desal = re.compile(r"(c[\-\_ ]?\d{1,3})", flags=re.IGNORECASE)
                
                for n in nombres_col:
                    m = patron_desal.search(str(n))
                    if m:
                        desaladores_detectados.add(re.sub(r"[\-_\s]", "", m.group(1)).upper())
                
                desaladores_detectados = sorted(list(desaladores_detectados))
                
                # --- LÃ³gica automÃ¡tica ---
                if len(desaladores_detectados) == 0:
                    desal_sel = ["GENERAL"]
                elif len(desaladores_detectados) == 1:
                    desal_sel = ["GENERAL"]
                else:
                    # varios desaladores â permitir selecciÃ³n en sidebar
                    st.sidebar.markdown("---")
                    st.sidebar.info(f"Detectados varios desaladores: {', '.join(desaladores_detectados)}")
                    desal_sel = st.sidebar.multiselect(
                        "Selecciona desaladores a analizar",
                        desaladores_detectados,
                        default=desaladores_detectados
                    )

                # ===== ConstrucciÃ³n REAL de variables base (igual que tu programa principal) =====
                mapa_variable_keys = list(mapa_variable_a_columnas.keys()) if mapa_variable_a_columnas else []
                variables_base = list(mapa_variable_keys)

                variables_base_limpias = []
                for v in variables_base:
                    partes = v.split()
                    if len(partes) > 1 and partes[-1].upper().startswith("C") and partes[-1][1:].isdigit():
                        variables_base_limpias.append(" ".join(partes[:-1]))
                    else:
                        variables_base_limpias.append(v)

                # Eliminar duplicados manteniendo orden
                variables_base_limpias = list(dict.fromkeys(variables_base_limpias))

                # Quitar tokens que parezcan fechas
                variables_base_filtradas = [v for v in variables_base_limpias if not es_token_fecha_like(clean_token(v))]

                if len(variables_base_filtradas) == 0:
                    # Si no quedan, usar limpias originales
                    variables_base_filtradas = variables_base_limpias

                # Determinar opciones a mostrar en el desplegable
                if len(desaladores_detectados) > 1:
                    st.sidebar.info(f"Se detectan varios desaladores: {', '.join(desaladores_detectados)}. Mostrando nombres base simples.")
                    opciones_variables_base = variables_base_filtradas
                else:
                    st.sidebar.info("Se detecta 1 desalador (o ninguno); se mostraran nombres completos.")
                    # en caso de 1 desalador mostramos todos los nombres tal cual (como en tu programa original)
                    opciones_variables_base = nombres_col if nombres_col else variables_base_filtradas

                # eliminar duplicados y ordenar manteniendo orden original
                opciones_variables_base = list(dict.fromkeys(opciones_variables_base))

                if not opciones_variables_base:
                    st.error("No se pudieron construir las opciones de variables base. Revisa el encabezado del Excel.")
                else:
                    var_sel = st.selectbox("Selecciona variable base", options=opciones_variables_base)

                    # MultiselecciÃ³n de desaladores (si hay varios)
                    if len(desaladores_detectados) > 1:
                        desal_sel = st.multiselect("Selecciona desaladores (filtrar)", options=desaladores_detectados, default=desaladores_detectados)
                    else:
                        desal_sel = st.multiselect("Selecciona desaladores (opcional)", options=desaladores_detectados, default=desaladores_detectados)

                    # Si el usuario escoge el nombre bÃ¡sico (cuando hay varios desaladores), tenemos que mapearlo a las columnas completas
                    # Construir cols_relacionadas a partir de mapa_variable_a_columnas
                    cols_relacionadas = []
                    if mapa_variable_a_columnas and var_sel in mapa_variable_a_columnas:
                        cols_relacionadas = mapa_variable_a_columnas[var_sel]
                        # si el usuario ha filtrado desaladores, aplicar filtro
                        if desal_sel:
                            filtered = []
                            for c in cols_relacionadas:
                                last_token = str(c).strip().split()[-1]
                                if any(d.upper() == re.sub(r"[\-_\s]", "", last_token).upper() for d in desal_sel):
                                    filtered.append(c)
                            if filtered:
                                cols_relacionadas = filtered
                    else:
                        # fallback: buscar columnas que contengan el string var_sel
                        if var_sel in datos.columns:
                            cols_relacionadas = [var_sel]
                        else:
                            cols_relacionadas = [c for c in datos.columns if var_sel.lower() in c.lower()]

                    st.write(f"Columnas relacionadas con '{var_sel}': {cols_relacionadas}")

                    st.markdown("---")
                    colA, colB = st.columns(2)

                    with colA:
                        valor_critico = st.number_input('Valor crÃ­tico (para análisis)', value=0.0, format="%.6f")
                        if st.button('Ejecutar análisis crítico'):
                            out_dir = Path.cwd() / 'Resultados_Desalacion_App' / 'Analisis_Criticos'
                            out_dir.mkdir(parents=True, exist_ok=True)
                            try:
                                archivos = analisis_critico_extendido(datos, desal_sel or list(desaladores_detectados or []), var_sel, float(valor_critico), str(out_dir), mapa_norm_columns)
                                st.success(f'Análisis crí­tico generado. Archivos: {len(archivos)}')
                                for k,v in archivos.items():
                                    try:
                                        with open(v, "rb") as f:
                                            st.download_button(f"Descargar {Path(v).name}", data=f, file_name=Path(v).name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                                    except Exception as e:
                                        st.write(f"No se pudo preparar descarga para {v}: {e}")
                            except Exception as e:
                                st.error(f'Error generando análisis crí­tico: {e}')

                    with colB:
                        if st.button('Generar gráficas por desalador'):
                            out_dir = Path.cwd() / 'Resultados_Desalacion_App' / 'Graficas'
                            out_dir.mkdir(parents=True, exist_ok=True)
                            try:
                                archivos_g = generar_graficas_por_desalador(datos, desal_sel or list(desaladores_detectados or []), var_sel, str(out_dir), mapa_norm_columns)
                                st.success(f'Gráficas generadas. Archivos: {len(archivos_g)}')
                                for k,v in archivos_g.items():
                                    try:
                                        with open(v, "rb") as f:
                                            st.download_button(f"Descargar {Path(v).name}", data=f, file_name=Path(v).name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                                    except Exception as e:
                                        st.write(f"No se pudo preparar descarga para {v}: {e}")
                            except Exception as e:
                                st.error(f'Error generando gráficas: {e}')

                    st.markdown("---")
                    st.subheader('Visualizaciones interactivas')
                    try:
                        cols_plot = [c for c in datos.columns if c != 'Tiempo']
                        if not cols_plot:
                            st.info("No hay columnas numÃ©ricas para graficar.")
                        else:
                            ycol = st.selectbox('Variable a graficar', options=cols_plot, index=0)
                            xmode = st.radio('Eje X', ['Tiempo','Variable base'], index=0)
                            fig, ax = plt.subplots(figsize=(fig_w, fig_h))
                            if xmode == 'Tiempo':
                                ax.scatter(pd.to_datetime(datos['Tiempo']), datos[ycol], s=10, alpha=0.7)
                                ax.set_xlabel('Tiempo')
                            else:
                                try:
                                    if len(cols_relacionadas) == 1:
                                        xseries = datos[cols_relacionadas[0]]
                                    else:
                                        xseries = datos[cols_relacionadas].mean(axis=1, skipna=True)
                                    ax.scatter(xseries, datos[ycol], s=10, alpha=0.7)
                                    ax.set_xlabel(var_sel)
                                except Exception:
                                    ax.scatter(datos.index, datos[ycol], s=10, alpha=0.7)
                                    ax.set_xlabel('Index')
                            ax.set_ylabel(ycol)
                            ax.grid(True)
                            st.pyplot(fig)
                    except Exception as e:
                        st.error(f'Error dibujando visualizaciÃ³n: {e}')

                    st.markdown("---")
                    st.subheader("Exportar datos procesados")
                    try:
                        to_export = datos.copy()
                        tmpfile = Path(tempfile.gettempdir()) / "datos_procesados_desalacion.xlsx"
                        to_export.to_excel(tmpfile, index=False)
                        with open(tmpfile, "rb") as f:
                            st.download_button("Descargar datos procesados (Excel)", data=f, file_name="datos_procesados_desalacion.xlsx",
                                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    except Exception as e:
                        st.error(f"No se pudo preparar exportaciÃ³n: {e}")

        else:
            st.info("El archivo no contiene hojas válidas o no se pudo leer.")

    st.markdown("---")

    # FIN DEL ARCHIVO

# =============================================================
# 2) PESTAÑA ANÁLISIS (AMPLIADA: modelos, SHAP, LIME, exclusión)
# =============================================================
with tab_rf:
    st.header("Análisis de importancia y explicaciones (RandomForest / SHAP / LIME)")
    st.write("Selecciona objetivo, desalador y modelo. Puedes excluir variables antes del entrenamiento.")

    if "datos" not in st.session_state:
        st.warning("Primero carga un archivo en la pestaña *Graficado*. Luego vuelve aquí.")
    else:
        datos = st.session_state["datos"].copy()

        # columnas numéricas disponibles (sin Tiempo)
        columnas_num = [c for c in datos.columns if c != "Tiempo"]

        # Detectar desaladores presentes por token C##
        import re
        patron_desal = re.compile(r"(c[\-\_ ]?\d{1,3})", flags=re.IGNORECASE)
        desaladores_detectados = set()
        for n in datos.columns:
            m = patron_desal.search(str(n))
            if m:
                desaladores_detectados.add(re.sub(r"[\-_\s]", "", m.group(1)).upper())
        desaladores_detectados = sorted(list(desaladores_detectados))
        if not desaladores_detectados:
            desaladores_detectados = []

        # Selector desalador: GENERAL + detectados
        opciones_desal = ["GENERAL"] + desaladores_detectados
        desal_sel = st.selectbox("Desalador para análisis", options=opciones_desal, index=0)

        # Filtrar columnas según desalador seleccionado
        def columnas_para_desalador(df_cols, desal_sel):
            if desal_sel == "GENERAL":
                return [c for c in df_cols if c != "Tiempo"]
            token = desal_sel.lower()
            cols = []
            comunes = []
            for c in df_cols:
                if c == "Tiempo":
                    continue
                if token in str(c).lower().replace(" ", "").replace("-", "").replace("_", ""):
                    cols.append(c)
                else:
                    # consideramos comunes si no contienen token
                    comunes.append(c)
            # devolvemos cols seleccionadas + comunes para permitir uso conjunto
            return sorted(list(set(cols + comunes)), key=lambda x: str(x))
        columnas_disponibles = columnas_para_desalador(datos.columns, desal_sel)

        # Selección de variable objetivo (Y)
        if not columnas_disponibles:
            st.error("No hay columnas disponibles para análisis.")
        else:
            target = st.selectbox("Variable objetivo (Y)", columnas_disponibles)

            # Multiselect para excluir variables
            excluir = st.multiselect("Variables a excluir del análisis (opcional)", options=[c for c in columnas_disponibles if c != target])

            # Construcción de X,y aplicando exclusión y eliminando NaNs en target
            df_model = datos.drop(columns=["Tiempo"]).copy()
            df_model = df_model.dropna(subset=[target])
            # eliminar columnas excluidas y la propia target de X
            X = df_model.drop(columns=[target] + excluir, errors='ignore')
            y = df_model[target].astype(float)

            if X.shape[0] < 10 or X.shape[1] == 0:
                st.error("Hay muy pocos datos válidos o no hay variables para entrenar el modelo.")
            else:
                # Selector de modelo
                modelo_sel = st.selectbox("Modelo", [
                    "Random Forest",
                    "Gradient Boosting",
                    "XGBoost",
                    "LightGBM",
                    "Ridge",
                    "Lasso"
                ])

                # Parámetros básicos
                n_estim = st.slider("Nº árboles (si aplica)", 50, 1000, 200)
                max_depth = st.slider("Max depth (si aplica, 0=auto)", 0, 30, 6)
                usar_shap = st.checkbox("Calcular SHAP (explicación global y local)", value=True)
                usar_lime = st.checkbox("Habilitar LIME (explicación local por instancia)", value=False)

                # Entrenar modelo según selección
                model = None
                if modelo_sel == "Random Forest":
                    from sklearn.ensemble import RandomForestRegressor
                    model = RandomForestRegressor(n_estimators=n_estim, max_depth=(None if max_depth==0 else max_depth), random_state=42, n_jobs=-1)
                elif modelo_sel == "Gradient Boosting":
                    model = GradientBoostingRegressor(n_estimators=n_estim, max_depth=(None if max_depth==0 else max_depth), random_state=42)
                elif modelo_sel == "XGBoost":
                    model = xgb.XGBRegressor(n_estimators=n_estim, max_depth=(0 if max_depth==0 else max_depth), random_state=42, n_jobs=-1)
                elif modelo_sel == "LightGBM":
                    model = lgb.LGBMRegressor(n_estimators=n_estim, max_depth=( -1 if max_depth==0 else max_depth), random_state=42, n_jobs=-1)
                elif modelo_sel == "Ridge":
                    model = Ridge()
                elif modelo_sel == "Lasso":
                    model = Lasso()

                # Entrenamiento
                with st.spinner("Entrenando modelo..."):
                    model.fit(X.fillna(0), y)  # rellenamos NaNs para entrenamiento simple

                st.success("Modelo entrenado")
                # Importancias simples (si el modelo tiene atributo feature_importances_)
                if hasattr(model, "feature_importances_"):
                    importancias = model.feature_importances_
                    ranking = pd.DataFrame({"Variable": X.columns, "Importancia": importancias}).sort_values("Importancia", ascending=False)
                    st.subheader("Ranking por importancia (modelo)")
                    st.dataframe(ranking)
                    fig, ax = plt.subplots(figsize=(8, max(4, 0.2*len(ranking))))
                    ax.barh(ranking["Variable"], ranking["Importancia"])
                    ax.invert_yaxis()
                    ax.set_xlabel("Importancia")
                    ax.set_ylabel("Variable")
                    st.pyplot(fig)
                else:
                    st.info("El modelo seleccionado no proporciona 'feature_importances_' (p.ej. Ridge/Lasso). Usar SHAP para explicaciones detalladas.")

                # ===============================
                # SHAP EN STREAMLIT (TOTALMENTE COMPATIBLE)
                # ===============================
                if usar_shap:
                    st.markdown("---")
                    st.subheader("SHAP - explicaciones globales y locales (compatibles con Streamlit)")
                
                    try:
                        # 1) Crear explainer según modelo
                        if modelo_sel in ["Random Forest", "Gradient Boosting", "XGBoost", "LightGBM"]:
                            explainer = shap.TreeExplainer(model)
                            shap_values = explainer.shap_values(X.fillna(0))
                        else:
                            explainer = shap.LinearExplainer(model, X.fillna(0), feature_perturbation="interventional")
                            shap_values = explainer.shap_values(X.fillna(0))
                
                        # Convertir shap_values a array si viene como lista
                        import numpy as np
                        if isinstance(shap_values, list):
                            shap_values = shap_values[0]
                
                        # 2) SHAP Summary (bar)
                        st.write("### SHAP Summary (bar)")
                        fig = plt.figure(figsize=(8,5))
                        shap.summary_plot(shap_values, X, plot_type="bar", show=False)
                        st.pyplot(fig)
                        plt.close(fig)
                
                        # 3) SHAP Beeswarm
                        st.write("### SHAP Beeswarm")
                        fig = plt.figure(figsize=(8,6))
                        shap.summary_plot(shap_values, X, show=False)
                        st.pyplot(fig)
                        plt.close(fig)
                
                        # 4) SHAP Dependence plot
                        col_dependence = st.selectbox("Variable para dependence plot", list(X.columns))
                        fig = plt.figure(figsize=(8,6))
                        shap.dependence_plot(
                            col_dependence,
                            shap_values,
                            X,
                            ax=plt.gca(),
                            show=False
                        )
                        st.pyplot(fig)
                        plt.close(fig)
                
                        # 5) Force plot (versión matplotlib, 100% compatible)
                        st.write("### SHAP Force Plot (local)")
                        idx_inst = st.number_input("Índice de la instancia", 0, len(X)-1, 0)
                
                        fig = shap.plots.force(
                            explainer.expected_value,
                            shap_values[idx_inst],
                            matplotlib=True
                        )
                        st.pyplot(fig)
                        plt.close(fig)
                
                    except Exception as e:
                        st.error(f"Error calculando SHAP: {e}")


                # ---------------------------
                # LIME: explicación local para una instancia
                # ---------------------------
                if usar_lime:
                    st.markdown("---")
                    st.subheader("LIME - explicación local (instancia)")
                    try:
                        i_inst = st.number_input("Índice para explicación LIME (0 .. n-1)", min_value=0, max_value=max(0, len(X)-1), value=0, step=1)
                        explainer_lime = LimeTabularExplainer(training_data=X.fillna(0).values,
                                                              feature_names=list(X.columns),
                                                              mode='regression')
                        exp = explainer_lime.explain_instance(X.fillna(0).iloc[i_inst].values, model.predict, num_features=min(10, X.shape[1]))
                        html = exp.as_html()
                        components.html(html, height=400)
                    except Exception as e:
                        st.error(f"Error con LIME: {e}")

                # ---------------------------
                # Exportar ranking (si existe) y SHAP values opcional
                # ---------------------------
                st.markdown("---")
                st.subheader("Exportar resultados")
                if 'ranking' in locals():
                    csv = ranking.to_csv(index=False).encode('utf-8')
                    st.download_button("Descargar CSV de importancias", data=csv, file_name="importancias_modelo.csv", mime="text/csv")
                if usar_shap:
                    try:
                        # preparar shap_values para exportar: convertir a DataFrame promedio absoluto
                        import numpy as np
                        mean_abs_shap = np.abs(shap_values).mean(axis=0)
                        shap_df = pd.DataFrame({"Variable": X.columns, "MeanAbsSHAP": mean_abs_shap}).sort_values("MeanAbsSHAP", ascending=False)
                        csv2 = shap_df.to_csv(index=False).encode('utf-8')
                        st.download_button("Descargar resumen SHAP (Mean abs)", data=csv2, file_name="shap_resumen.csv", mime="text/csv")
                    except Exception:
                        st.info("No se pudo exportar SHAP (problema al calcular shap_values).")

